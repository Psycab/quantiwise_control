"""
DeepSearch 외인수급Top20 지수 (PR) 매달 리밸런싱 시스템
영문명: DeepSearch Net Foreign BuyingTop20 Index PR

핵심 로직:
1. EPS 필터: (1개월 평균 - 3개월 평균) / abs(3개월 평균) → 상위 100개
2. 외국인 수급강도: 6개월 외국인 순매수 평균 / 6개월 시가총액 평균 → 상위 50개
3. 월별 상위 10개 선정 및 최종 비중 계산
4. 매달 리밸런싱 자동화 (파일 복사, 날짜 업데이트, Excel 자동화)
"""

import numpy as np
import openpyxl
from openpyxl import load_workbook, Workbook
import os
import shutil
from datetime import datetime
import calendar
import time

class DeepSearchForeignBuyingTop20IndexSystem:
    """DeepSearch 외인수급Top20 지수 분석 시스템"""
    
    def __init__(self, source_excel_path, output_excel_path):
        self.source_excel_path = source_excel_path
        self.output_excel_path = output_excel_path
        self.source_workbook = None
        self.output_workbook = None
        
    def load_source_excel_file(self):
        """소스 Excel 파일 로드"""
        try:
            # UTF-8 인코딩으로 Excel 파일 로드
            self.source_workbook = load_workbook(self.source_excel_path, data_only=True)
            print(f"소스 Excel 파일 로드 완료")
            return True
        except Exception as e:
            print(f"소스 Excel 파일 로드 실패: {e}")
            return False
    
    def find_data_sheets(self, use_market_cap=True):
        """데이터 시트 찾기"""
        sheets = {}
        for sheet_name in self.source_workbook.sheetnames:
            if "eps" in sheet_name:
                sheets['eps_sheet'] = sheet_name
            elif "foreign" in sheet_name:
                sheets['foreign_sheet'] = sheet_name
            elif use_market_cap and "market_cap" in sheet_name:
                sheets['market_cap_sheet'] = sheet_name
            elif not use_market_cap and "market_ff_cap" in sheet_name:
                sheets['market_cap_sheet'] = sheet_name
        
        print(f"발견된 시트: {list(sheets.keys())}")
        return sheets
    
    def parse_data(self, sheet_name, data_type):
        """데이터 파싱"""
        try:
            print(f"{data_type} 데이터 파싱 중...")
            
            worksheet = self.source_workbook[sheet_name]
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # 종목코드와 종목명 추출 (8행, 9행)
            stock_codes = []
            stock_names = {}
            
            for col in range(2, max_col + 1):
                try:
                    code_value = worksheet.cell(row=8, column=col).value
                    name_value = worksheet.cell(row=9, column=col).value
                    
                    if code_value and str(code_value).strip():
                        try:
                            clean_code = str(code_value).strip()
                            stock_codes.append(clean_code)
                            if name_value:
                                clean_name = str(name_value).strip()
                                stock_names[clean_code] = clean_name
                        except UnicodeError:
                            # UTF-8 인코딩 오류 시 기본값 사용
                            clean_code = f"종목_{col-2}"
                            stock_codes.append(clean_code)
                            stock_names[clean_code] = f"종목명_{col-2}"
                except:
                    continue
            
            print(f"종목코드 추출 완료: {len(stock_codes)}개")
            
            # 시계열 데이터 추출 (15행부터 시작, DATE 헤더는 14행)
            data = {}
            start_row = 15  # 실제 데이터 시작 행 (DATE 헤더 다음)
            
            # 데이터가 있는 마지막 행을 동적으로 찾기
            end_row = start_row
            for row in range(start_row, max_row + 1):
                date_cell = worksheet.cell(row=row, column=1).value
                if date_cell is not None and isinstance(date_cell, datetime):
                    end_row = row
                elif date_cell is None:
                    break
            
            print(f"  [정보] 데이터 범위: A{start_row} ~ A{end_row} (총 {end_row - start_row + 1}행)")
            
            # 날짜 정보 추출 (15행부터의 날짜들)
            dates = []
            for row in range(start_row, end_row + 1):
                date_cell = worksheet.cell(row=row, column=1).value  # A열의 날짜
                if date_cell is not None:
                    try:
                        # 날짜 형식 변환 (Excel 날짜 또는 문자열)
                        if isinstance(date_cell, datetime):
                            dates.append(date_cell)
                        else:
                            # 문자열인 경우 다양한 형식으로 변환 시도
                            date_str = str(date_cell).strip()
                            
                            # 빈 문자열이나 비정상적인 문자열 제외
                            if date_str.upper() in ['DATE', '날짜', ''] or len(date_str) < 4:
                                continue
                            
                            # 숫자만 있는 경우 (YYYYMMDD 형식)
                            if date_str.isdigit() and len(date_str) == 8:
                                dates.append(datetime.strptime(date_str, '%Y%m%d'))
                            # 하이픈이나 슬래시가 있는 경우
                            elif '-' in date_str or '/' in date_str:
                                clean_str = date_str.replace('-', '').replace('/', '').replace(' ', '')
                                if clean_str.isdigit() and len(clean_str) == 8:
                                    dates.append(datetime.strptime(clean_str, '%Y%m%d'))
                                else:
                                    # 다른 형식 시도
                                    try:
                                        dates.append(datetime.strptime(date_str, '%Y-%m-%d'))
                                    except:
                                        try:
                                            dates.append(datetime.strptime(date_str, '%Y/%m/%d'))
                                        except:
                                            continue
                            else:
                                continue
                    except Exception as e:
                        # 변환 실패 시 해당 행 건너뛰기
                        continue
            
            for i, stock_code in enumerate(stock_codes):
                col = i + 2
                values = []
                valid_indices = []  # 유효한 데이터의 인덱스 저장
                
                for row in range(start_row, end_row + 1):
                    try:
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value is not None:
                            if data_type == "foreign":
                                value = float(cell_value) * 100000000  # 외국인 순매수는 10000 곱하기
                            else:
                                value = float(cell_value)
                            values.append(value)
                            valid_indices.append(row - start_row)  # 0부터 시작하는 인덱스
                        # 빈 셀은 제외 (0을 추가하지 않음)
                    except:
                        # 변환 실패한 경우도 제외
                        pass
                
                if values:
                    # 날짜와 값의 개수를 맞춤 (유효한 데이터만)
                    valid_dates = []
                    for idx in valid_indices:
                        if idx < len(dates):
                            valid_dates.append(dates[idx])
                    
                    data[stock_code] = {
                        'name': stock_names.get(stock_code, f"종목_{stock_code}"),
                        'values': np.array(values),
                        'dates': valid_dates,
                        'valid_indices': valid_indices  # 유효한 데이터의 인덱스 정보 추가
                    }
            
            print(f"{data_type} 데이터 추출 완료: {len(data)}개 종목 (전체 {len(stock_codes)}개 중)")
            
            # 데이터 날짜 범위 출력
            if dates:
                try:
                    start_date = dates[0] if isinstance(dates[0], datetime) else datetime.strptime(str(dates[0]), '%Y%m%d')
                    end_date = dates[-1] if isinstance(dates[-1], datetime) else datetime.strptime(str(dates[-1]), '%Y%m%d')
                    print(f"  [날짜] {data_type} 데이터 기간: {start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')} ({len(dates)}일)")
                except:
                    print(f"  [경고] {data_type} 데이터: 날짜 형식 오류")
            else:
                print(f"  [경고] {data_type} 데이터: 날짜 정보 없음")
            
            return data, len(stock_codes)  # 데이터와 전체 종목 수 반환
            
        except Exception as e:
            print(f"[오류] {data_type} 데이터 파싱 실패: {e}")
            return None, 0
    
    def apply_eps_filter(self, eps_data):
        """EPS 필터 적용: (1개월 평균 - 3개월 평균) / abs(3개월 평균)"""
        try:
            print("EPS 필터 적용 중...")
            
            eps_scores = {}
            
            for stock_code, data in eps_data.items():
                eps_values = data.get('values', np.array([]))
                dates = data.get('dates', [])
                
                if len(eps_values) < 30:
                    eps_scores[stock_code] = {
                        'name': data.get('name', f"종목_{stock_code}"),
                        'eps_score': 0,
                        'status': '데이터부족'
                    }
                    continue
                
                # B6 날짜 기준으로 정확한 기간 계산
                # 1개월 EPS 평균: B6 기준 1개월 전부터 B6까지
                # 3개월 EPS 평균: B6 기준 3개월 전부터 B6까지
                
                # 날짜가 있는 경우 날짜 기반으로 계산, 없으면 개수 기반으로 계산
                if dates and len(dates) == len(eps_values):
                    # 날짜 기반 계산 (정확한 방식)
                    from datetime import datetime, timedelta
                    import calendar
                    
                    def get_month_start_date(target_date, months_back):
                        """정확한 월 단위 계산 - N개월 전부터 현재까지"""
                        year = target_date.year
                        month = target_date.month
                        
                        # N개월 전 월 계산 (포함)
                        month -= (months_back - 1)
                        while month <= 0:
                            month += 12
                            year -= 1
                        
                        # 해당 월의 첫 번째 날
                        return datetime(year, month, 1)
                    
                    # B6 날짜 (최종 날짜) - 실제로는 데이터의 마지막 날짜 사용
                    end_date = dates[-1] if isinstance(dates[-1], datetime) else datetime.strptime(str(dates[-1]), '%Y%m%d')
                    
                    # 1개월 전 첫째 날
                    one_month_start = get_month_start_date(end_date, 1)
                    # 3개월 전 첫째 날
                    three_month_start = get_month_start_date(end_date, 3)
                    
                    # 1개월 데이터 추출
                    one_month_values = []
                    three_month_values = []
                    
                    for i, date in enumerate(dates):
                        if isinstance(date, str):
                            date = datetime.strptime(date, '%Y%m%d')
                        
                        if date >= one_month_start:
                            one_month_values.append(eps_values[i])
                        if date >= three_month_start:
                            three_month_values.append(eps_values[i])
                    
                    # 빈 셀은 이미 제외된 상태이므로 실제 데이터만으로 평균 계산
                    one_month_avg = np.mean(one_month_values) if one_month_values else 0
                    three_month_avg = np.mean(three_month_values) if three_month_values else 0
                    
                    # 첫 번째 종목에서만 날짜 범위 출력
                    if stock_code == list(eps_data.keys())[0]:
                        print(f"  [날짜] EPS 필터 계산 기간:")
                        print(f"     - 1개월 평균: {one_month_start.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')} ({len(one_month_values)}일)")
                        print(f"     - 3개월 평균: {three_month_start.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')} ({len(three_month_values)}일)")
                    
                else:
                    # 날짜가 없는 경우 개수 기반 계산 (기존 방식)
                    one_month_avg = np.mean(eps_values[-30:])
                    three_month_avg = np.mean(eps_values[-90:]) if len(eps_values) >= 90 else np.mean(eps_values)
                    
                    # 첫 번째 종목에서만 정보 출력
                    if stock_code == list(eps_data.keys())[0]:
                        print(f"  [경고] EPS 필터: 날짜 정보 없음, 개수 기반 계산 (최근 30일, 90일)")
                
                # EPS 점수 계산
                if abs(three_month_avg) > 1e-6:
                    eps_score = (one_month_avg - three_month_avg) / abs(three_month_avg)
                else:
                    eps_score = 0
                
                eps_scores[stock_code] = {
                    'name': data.get('name', f"종목_{stock_code}"),
                    'eps_score': eps_score,
                    'one_month_avg': one_month_avg,
                    'three_month_avg': three_month_avg,
                    'status': '계산완료'
                }
            
            # EPS 점수 기준으로 정렬하여 상위 100개 선정
            sorted_stocks = sorted(eps_scores.items(), key=lambda x: x[1]['eps_score'], reverse=True)
            top_100_stocks = dict(sorted_stocks[:100])
            
            print(f"EPS 필터 적용 완료: 전체 {len(eps_scores)}개 종목 중 상위 100개 선정")
            
            self.eps_scores = eps_scores
            self.eps_top_100 = top_100_stocks
            
            return top_100_stocks
            
        except Exception as e:
            print(f"EPS 필터 적용 실패: {e}")
            return None
    
    def calculate_foreign_intensity(self, eps_filtered_stocks, foreign_data, market_cap_data):
        """외국인 수급강도 지표 계산: 6개월 외국인 순매수 평균 / 6개월 시가총액 평균"""
        try:
            print("외국인 수급강도 지표 계산 중...")
            
            intensity_scores = {}
            
            # EPS 필터를 통과한 종목들만 처리
            for stock_code, eps_data in eps_filtered_stocks.items():
                if stock_code not in foreign_data or stock_code not in market_cap_data:
                    intensity_scores[stock_code] = {
                        'name': eps_data.get('name', f"종목_{stock_code}"),
                        'intensity_score': 0,
                        'status': '데이터부족'
                    }
                    continue
                
                foreign_values = foreign_data[stock_code].get('values', np.array([]))
                cap_values = market_cap_data[stock_code].get('values', np.array([]))
                foreign_dates = foreign_data[stock_code].get('dates', [])
                cap_dates = market_cap_data[stock_code].get('dates', [])
                
                if len(foreign_values) < 30 or len(cap_values) < 30:
                    intensity_scores[stock_code] = {
                        'name': eps_data.get('name', f"종목_{stock_code}"),
                        'intensity_score': 0,
                        'status': '데이터부족'
                    }
                    continue
                
                # 6개월 평균 계산 - 날짜 기반으로 정확한 기간 계산
                if foreign_dates and cap_dates and len(foreign_dates) == len(foreign_values) and len(cap_dates) == len(cap_values):
                    # 날짜 기반 계산 (정확한 방식)
                    from datetime import datetime, timedelta
                    
                    def get_month_start_date(target_date, months_back):
                        """정확한 월 단위 계산 - N개월 전부터 현재까지"""
                        year = target_date.year
                        month = target_date.month
                        
                        # N개월 전 월 계산 (포함)
                        month -= (months_back - 1)
                        while month <= 0:
                            month += 12
                            year -= 1
                        
                        # 해당 월의 첫 번째 날
                        return datetime(year, month, 1)
                    
                    # 최종 날짜 (데이터의 마지막 날짜)
                    end_date = foreign_dates[-1] if isinstance(foreign_dates[-1], datetime) else datetime.strptime(str(foreign_dates[-1]), '%Y%m%d')
                    
                    # 6개월 전 첫째 날
                    six_month_start = get_month_start_date(end_date, 6)
                    
                    # 6개월 데이터 추출
                    foreign_6month_values = []
                    cap_6month_values = []
                    
                    for i, date in enumerate(foreign_dates):
                        if isinstance(date, str):
                            date = datetime.strptime(date, '%Y%m%d')
                        
                        if date >= six_month_start:
                            foreign_6month_values.append(foreign_values[i])
                            if i < len(cap_values):
                                cap_6month_values.append(cap_values[i])
                    
                    # 빈 셀은 이미 제외된 상태이므로 실제 데이터만으로 평균 계산
                    foreign_avg = np.mean(foreign_6month_values) if foreign_6month_values else 0
                    cap_avg = np.mean(cap_6month_values) if cap_6month_values else 0
                    
                    # 첫 번째 종목에서만 날짜 범위 출력
                    if stock_code == list(eps_filtered_stocks.keys())[0]:
                        print(f"  [날짜] 외국인 수급강도 계산 기간:")
                        print(f"     - 6개월 평균: {six_month_start.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')} ({len(foreign_6month_values)}일)")
                    
                else:
                    # 날짜가 없는 경우 개수 기반 계산 (기존 방식)
                    foreign_avg = np.mean(foreign_values[-min(180, len(foreign_values)):])
                    cap_avg = np.mean(cap_values[-min(180, len(cap_values)):])
                    
                    # 첫 번째 종목에서만 정보 출력
                    if stock_code == list(eps_filtered_stocks.keys())[0]:
                        print(f"  [경고] 외국인 수급강도: 날짜 정보 없음, 개수 기반 계산 (최근 180일)")
                
                # 외국인 수급강도 지표 계산
                if cap_avg > 1e-6:
                    intensity_score = foreign_avg / cap_avg
                else:
                    intensity_score = 0
                
                intensity_scores[stock_code] = {
                    'name': eps_data.get('name', f"종목_{stock_code}"),
                    'intensity_score': intensity_score,
                    'foreign_avg': foreign_avg,
                    'cap_avg': cap_avg,
                    'eps_score': eps_data.get('eps_score', 0),
                    'status': '계산완료'
                }
            
            # 외국인 수급강도 지표 기준으로 정렬하여 상위 50개 선정
            sorted_stocks = sorted(intensity_scores.items(), key=lambda x: x[1]['intensity_score'], reverse=True)
            top_50_stocks = dict(sorted_stocks[:50])
            
            print(f"외국인 수급강도 지표 계산 완료: 상위 50개 종목 선정")
            
            self.intensity_scores = intensity_scores
            self.final_top_50 = top_50_stocks
            
            return top_50_stocks
            
        except Exception as e:
            print(f"외국인 수급강도 지표 계산 실패: {e}")
            return None
    
    def calculate_monthly_foreign_intensity(self, final_stocks, foreign_data, market_cap_data):
        """1개월과 2개월 외국인 수급 상위 10종목 계산"""
        try:
            print("1개월과 2개월 외국인 수급 상위 10종목 계산 중...")
            
            one_month_scores = {}
            two_month_scores = {}
            
            for stock_code, data in final_stocks.items():
                if stock_code not in foreign_data or stock_code not in market_cap_data:
                    continue
                
                foreign_values = foreign_data[stock_code].get('values', np.array([]))
                cap_values = market_cap_data[stock_code].get('values', np.array([]))
                foreign_dates = foreign_data[stock_code].get('dates', [])
                cap_dates = market_cap_data[stock_code].get('dates', [])
                
                if len(foreign_values) < 30 or len(cap_values) < 30:
                    continue
                
                # 1개월과 2개월 평균 계산 - 날짜 기반으로 정확한 기간 계산
                if foreign_dates and cap_dates and len(foreign_dates) == len(foreign_values) and len(cap_dates) == len(cap_values):
                    # 날짜 기반 계산 (정확한 방식)
                    from datetime import datetime, timedelta
                    
                    def get_month_start_date(target_date, months_back):
                        """정확한 월 단위 계산 - N개월 전부터 현재까지"""
                        year = target_date.year
                        month = target_date.month
                        
                        # N개월 전 월 계산 (포함)
                        month -= (months_back - 1)
                        while month <= 0:
                            month += 12
                            year -= 1
                        
                        # 해당 월의 첫 번째 날
                        return datetime(year, month, 1)
                    
                    # 최종 날짜 (데이터의 마지막 날짜)
                    end_date = foreign_dates[-1] if isinstance(foreign_dates[-1], datetime) else datetime.strptime(str(foreign_dates[-1]), '%Y%m%d')
                    
                    # 1개월 전 첫째 날
                    one_month_start = get_month_start_date(end_date, 1)
                    # 2개월 전 첫째 날
                    two_month_start = get_month_start_date(end_date, 2)
                    
                    # 1개월 데이터 추출
                    one_month_foreign_values = []
                    one_month_cap_values = []
                    # 2개월 데이터 추출
                    two_month_foreign_values = []
                    two_month_cap_values = []
                    
                    for i, date in enumerate(foreign_dates):
                        if isinstance(date, str):
                            date = datetime.strptime(date, '%Y%m%d')
                        
                        if date >= one_month_start:
                            one_month_foreign_values.append(foreign_values[i])
                            if i < len(cap_values):
                                one_month_cap_values.append(cap_values[i])
                        
                        if date >= two_month_start:
                            two_month_foreign_values.append(foreign_values[i])
                            if i < len(cap_values):
                                two_month_cap_values.append(cap_values[i])
                    
                    # 빈 셀은 이미 제외된 상태이므로 실제 데이터만으로 평균 계산
                    one_month_foreign = np.mean(one_month_foreign_values) if one_month_foreign_values else 0
                    one_month_cap = np.mean(one_month_cap_values) if one_month_cap_values else 0
                    two_month_foreign = np.mean(two_month_foreign_values) if two_month_foreign_values else 0
                    two_month_cap = np.mean(two_month_cap_values) if two_month_cap_values else 0
                    
                    # 첫 번째 종목에서만 날짜 범위 출력
                    if stock_code == list(final_stocks.keys())[0]:
                        print(f"  [날짜] 월별 외국인 수급 계산 기간:")
                        print(f"     - 1개월 평균: {one_month_start.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')} ({len(one_month_foreign_values)}일)")
                        print(f"     - 2개월 평균: {two_month_start.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')} ({len(two_month_foreign_values)}일)")
                    
                else:
                    # 날짜가 없는 경우 개수 기반 계산 (기존 방식)
                    one_month_foreign = np.mean(foreign_values[-30:])
                    one_month_cap = np.mean(cap_values[-30:])
                    two_month_foreign = np.mean(foreign_values[-60:]) if len(foreign_values) >= 60 else np.mean(foreign_values)
                    two_month_cap = np.mean(cap_values[-60:]) if len(cap_values) >= 60 else np.mean(cap_values)
                    
                    # 첫 번째 종목에서만 정보 출력
                    if stock_code == list(final_stocks.keys())[0]:
                        print(f"  [경고] 월별 외국인 수급: 날짜 정보 없음, 개수 기반 계산 (최근 30일, 60일)")
                
                # 1개월 외국인 수급 지표
                if one_month_cap > 1e-6:
                    one_month_score = one_month_foreign / one_month_cap
                else:
                    one_month_score = 0
                
                # 2개월 외국인 수급 지표
                if two_month_cap > 1e-6:
                    two_month_score = two_month_foreign / two_month_cap
                else:
                    two_month_score = 0
                
                one_month_scores[stock_code] = {
                    'name': data.get('name', f"종목_{stock_code}"),
                    'one_month_score': one_month_score,
                    'one_month_foreign': one_month_foreign,
                    'one_month_cap': one_month_cap,
                    'eps_score': data.get('eps_score', 0),
                    'intensity_score': data.get('intensity_score', 0)
                }
                
                two_month_scores[stock_code] = {
                    'name': data.get('name', f"종목_{stock_code}"),
                    'two_month_score': two_month_score,
                    'two_month_foreign': two_month_foreign,
                    'two_month_cap': two_month_cap,
                    'eps_score': data.get('eps_score', 0),
                    'intensity_score': data.get('intensity_score', 0)
                }
            
            # 1개월 상위 10개 선정
            sorted_one_month = sorted(one_month_scores.items(), key=lambda x: x[1]['one_month_score'], reverse=True)
            top_10_one_month = dict(sorted_one_month[:10])
            
            # 2개월 상위 10개 선정
            sorted_two_month = sorted(two_month_scores.items(), key=lambda x: x[1]['two_month_score'], reverse=True)
            top_10_two_month = dict(sorted_two_month[:10])
            
            print(f"1개월 외국인 수급 상위 10종목 선정 완료")
            print(f"2개월 외국인 수급 상위 10종목 선정 완료")
            
            # 결과 저장
            self.one_month_top_10 = top_10_one_month
            self.two_month_top_10 = top_10_two_month
            
            return top_10_one_month, top_10_two_month
            
        except Exception as e:
            print(f"월별 외국인 수급 계산 실패: {e}")
            return None, None
    
    def calculate_final_weights(self):
        """최종 비중 계산: (1개월과 2개월 선정 횟수) / (1개월과 2개월 선정 종목 개수 총합)"""
        try:
            print("최종 비중 계산 중...")
            
            # 종목별 선정 횟수 계산
            stock_selection_count = {}
            
            # 1개월 상위 10개 종목 카운트
            for stock_code in self.one_month_top_10.keys():
                if stock_code not in stock_selection_count:
                    stock_selection_count[stock_code] = 0
                stock_selection_count[stock_code] += 1
            
            # 2개월 상위 10개 종목 카운트
            for stock_code in self.two_month_top_10.keys():
                if stock_code not in stock_selection_count:
                    stock_selection_count[stock_code] = 0
                stock_selection_count[stock_code] += 1
            
            # 총 선정 종목 개수 계산 (1개월 + 2개월, 중복 포함)
            total_selection_count = len(self.one_month_top_10) + len(self.two_month_top_10)
            
            print(f"1개월 상위 10개 종목: {len(self.one_month_top_10)}개")
            print(f"2개월 상위 10개 종목: {len(self.two_month_top_10)}개")
            print(f"총 선정 종목 수 (중복 포함): {total_selection_count}개")
            
            # 최종 비중 계산
            final_weights = {}
            
            for stock_code, count in stock_selection_count.items():
                # 종목 정보 수집
                stock_name = None
                one_month_score = 0
                two_month_score = 0
                eps_score = 0
                intensity_score = 0
                
                # 1개월 정보
                if stock_code in self.one_month_top_10:
                    stock_name = self.one_month_top_10[stock_code].get('name')
                    one_month_rank = list(self.one_month_top_10.keys()).index(stock_code) + 1
                    one_month_score = self.one_month_top_10[stock_code].get('one_month_score', 0)
                    eps_score = self.one_month_top_10[stock_code].get('eps_score', 0)
                    intensity_score = self.one_month_top_10[stock_code].get('intensity_score', 0)
                else:
                    one_month_rank = None
                
                # 2개월 정보
                if stock_code in self.two_month_top_10:
                    if not stock_name:
                        stock_name = self.two_month_top_10[stock_code].get('name')
                    two_month_rank = list(self.two_month_top_10.keys()).index(stock_code) + 1
                    two_month_score = self.two_month_top_10[stock_code].get('two_month_score', 0)
                    # 1개월에서 가져오지 못한 점수 정보가 있다면 2개월에서 가져오기
                    if eps_score == 0:
                        eps_score = self.two_month_top_10[stock_code].get('eps_score', 0)
                    if intensity_score == 0:
                        intensity_score = self.two_month_top_10[stock_code].get('intensity_score', 0)
                else:
                    two_month_rank = None
                
                # 최종 비중 계산: (선정 횟수) / (총 선정 종목 수)
                final_weight = count / total_selection_count
                
                final_weights[stock_code] = {
                    'name': stock_name or f"종목_{stock_code}",
                    'selection_count': count,
                    'final_weight': final_weight,
                    'one_month_rank': one_month_rank,
                    'two_month_rank': two_month_rank,
                    'one_month_score': one_month_score,
                    'two_month_score': two_month_score,
                    'eps_score': eps_score,
                    'intensity_score': intensity_score
                }
            
            # 최종 비중 순으로 정렬
            sorted_final_weights = sorted(final_weights.items(), key=lambda x: x[1]['final_weight'], reverse=True)
            
            print(f"최종 비중 계산 완료: {len(final_weights)}개 종목")
            
            # 결과 저장
            self.final_weights = dict(sorted_final_weights)
            self.total_selection_count = total_selection_count
            
            return dict(sorted_final_weights)
            
        except Exception as e:
            print(f"최종 비중 계산 실패: {e}")
            return None
    
    def create_result_excel_full_stocks(self, final_stocks):
        """전체 종목 결과 엑셀 파일 생성"""
        try:
            print("전체 종목 결과 Excel 파일 생성 중...")
            
            # 새로운 워크북 생성
            self.output_workbook = Workbook()
            self.output_workbook.remove(self.output_workbook.active)
            
            # 1. 최종 결과 시트 (상위 50개)
            final_ws = self.output_workbook.create_sheet("최종구성종목50개")
            headers = ["순위", "종목코드", "종목명", "EPS점수", "외국인수급강도", "6개월외국인평균", "6개월시총평균", "상태"]
            for col, header in enumerate(headers, 1):
                final_ws.cell(row=1, column=col, value=header)
            
            row = 2
            for rank, (stock_code, data) in enumerate(final_stocks.items(), 1):
                final_ws.cell(row=row, column=1, value=rank)
                final_ws.cell(row=row, column=2, value=stock_code)
                final_ws.cell(row=row, column=3, value=data.get('name', f"종목_{stock_code}"))
                final_ws.cell(row=row, column=4, value=round(data.get('eps_score', 0), 4))
                final_ws.cell(row=row, column=5, value=round(data.get('intensity_score', 0), 6))
                final_ws.cell(row=row, column=6, value=round(data.get('foreign_avg', 0), 2))
                final_ws.cell(row=row, column=7, value=round(data.get('cap_avg', 0), 2))
                final_ws.cell(row=row, column=8, value=data.get('status', '알수없음'))
                row += 1
            
            # 2. EPS 필터 전체 결과 시트
            eps_ws = self.output_workbook.create_sheet("EPS필터전체결과")
            headers = ["순위", "종목코드", "종목명", "EPS점수", "1개월EPS평균", "3개월EPS평균", "데이터개수", "상태", "통과여부"]
            for col, header in enumerate(headers, 1):
                eps_ws.cell(row=1, column=col, value=header)
            
            row = 2
            sorted_eps = sorted(self.eps_scores.items(), key=lambda x: x[1]['eps_score'], reverse=True)
            for rank, (stock_code, data) in enumerate(sorted_eps, 1):
                eps_ws.cell(row=row, column=1, value=rank)
                eps_ws.cell(row=row, column=2, value=stock_code)
                eps_ws.cell(row=row, column=3, value=data.get('name', f"종목_{stock_code}"))
                eps_ws.cell(row=row, column=4, value=round(data.get('eps_score', 0), 4))
                eps_ws.cell(row=row, column=5, value=round(data.get('one_month_avg', 0), 2))
                eps_ws.cell(row=row, column=6, value=round(data.get('three_month_avg', 0), 2))
                eps_ws.cell(row=row, column=7, value=data.get('data_count', 0))
                eps_ws.cell(row=row, column=8, value=data.get('status', '알수없음'))
                eps_ws.cell(row=row, column=9, value="통과" if stock_code in self.eps_top_100 else "미통과")
                row += 1
            
            # 3. 외국인 수급강도 전체 결과 시트
            intensity_ws = self.output_workbook.create_sheet("외국인수급강도전체결과")
            headers = ["순위", "종목코드", "종목명", "수급강도지표", "6개월외국인평균", "6개월시총평균", "EPS점수", "상태", "통과여부"]
            for col, header in enumerate(headers, 1):
                intensity_ws.cell(row=1, column=col, value=header)
            
            row = 2
            sorted_intensity = sorted(self.intensity_scores.items(), key=lambda x: x[1]['intensity_score'], reverse=True)
            for rank, (stock_code, data) in enumerate(sorted_intensity, 1):
                intensity_ws.cell(row=row, column=1, value=rank)
                intensity_ws.cell(row=row, column=2, value=stock_code)
                intensity_ws.cell(row=row, column=3, value=data.get('name', f"종목_{stock_code}"))
                intensity_ws.cell(row=row, column=4, value=round(data.get('intensity_score', 0), 6))
                intensity_ws.cell(row=row, column=5, value=round(data.get('foreign_avg', 0), 2))
                intensity_ws.cell(row=row, column=6, value=round(data.get('cap_avg', 0), 2))
                intensity_ws.cell(row=row, column=7, value=round(data.get('eps_score', 0), 4))
                intensity_ws.cell(row=row, column=8, value=data.get('status', '알수없음'))
                intensity_ws.cell(row=row, column=9, value="통과" if stock_code in self.final_top_50 else "미통과")
                row += 1
            
            # 4. 1개월 외국인 수급 상위 10종목 시트
            one_month_ws = self.output_workbook.create_sheet("1개월외국인수급상위10개")
            headers = ["순위", "종목코드", "종목명", "1개월수급지표", "1개월외국인평균", "1개월시총평균", "EPS점수", "6개월수급지표"]
            for col, header in enumerate(headers, 1):
                one_month_ws.cell(row=1, column=col, value=header)
            
            row = 2
            for rank, (stock_code, data) in enumerate(self.one_month_top_10.items(), 1):
                one_month_ws.cell(row=row, column=1, value=rank)
                one_month_ws.cell(row=row, column=2, value=stock_code)
                one_month_ws.cell(row=row, column=3, value=data.get('name', f"종목_{stock_code}"))
                one_month_ws.cell(row=row, column=4, value=round(data.get('one_month_score', 0), 6))
                one_month_ws.cell(row=row, column=5, value=round(data.get('one_month_foreign', 0), 2))
                one_month_ws.cell(row=row, column=6, value=round(data.get('one_month_cap', 0), 2))
                one_month_ws.cell(row=row, column=7, value=round(data.get('eps_score', 0), 4))
                one_month_ws.cell(row=row, column=8, value=round(data.get('intensity_score', 0), 6))
                row += 1
            
            # 5. 2개월 외국인 수급 상위 10종목 시트
            two_month_ws = self.output_workbook.create_sheet("2개월외국인수급상위10개")
            headers = ["순위", "종목코드", "종목명", "2개월수급지표", "2개월외국인평균", "2개월시총평균", "EPS점수", "6개월수급지표"]
            for col, header in enumerate(headers, 1):
                two_month_ws.cell(row=1, column=col, value=header)
            
            row = 2
            for rank, (stock_code, data) in enumerate(self.two_month_top_10.items(), 1):
                two_month_ws.cell(row=row, column=1, value=rank)
                two_month_ws.cell(row=row, column=2, value=stock_code)
                two_month_ws.cell(row=row, column=3, value=data.get('name', f"종목_{stock_code}"))
                two_month_ws.cell(row=row, column=4, value=round(data.get('two_month_score', 0), 6))
                two_month_ws.cell(row=row, column=5, value=round(data.get('two_month_foreign', 0), 2))
                two_month_ws.cell(row=row, column=6, value=round(data.get('two_month_cap', 0), 2))
                two_month_ws.cell(row=row, column=7, value=round(data.get('eps_score', 0), 4))
                two_month_ws.cell(row=row, column=8, value=round(data.get('intensity_score', 0), 6))
                row += 1
            
            # 6. 최종 비중 시트
            final_weight_ws = self.output_workbook.create_sheet("최종비중순위")
            headers = ["순위", "종목코드", "종목명", "선정횟수", "최종비중", "1개월순위", "2개월순위", "1개월점수", "2개월점수", "EPS점수", "6개월수급지표"]
            for col, header in enumerate(headers, 1):
                final_weight_ws.cell(row=1, column=col, value=header)
            
            row = 2
            for rank, (stock_code, data) in enumerate(self.final_weights.items(), 1):
                final_weight_ws.cell(row=row, column=1, value=rank)
                final_weight_ws.cell(row=row, column=2, value=stock_code)
                final_weight_ws.cell(row=row, column=3, value=data.get('name', f"종목_{stock_code}"))
                final_weight_ws.cell(row=row, column=4, value=data.get('selection_count', 0))
                final_weight_ws.cell(row=row, column=5, value=round(data.get('final_weight', 0), 4))
                final_weight_ws.cell(row=row, column=6, value=data.get('one_month_rank', "-"))
                final_weight_ws.cell(row=row, column=7, value=data.get('two_month_rank', "-"))
                final_weight_ws.cell(row=row, column=8, value=round(data.get('one_month_score', 0), 6))
                final_weight_ws.cell(row=row, column=9, value=round(data.get('two_month_score', 0), 6))
                final_weight_ws.cell(row=row, column=10, value=round(data.get('eps_score', 0), 4))
                final_weight_ws.cell(row=row, column=11, value=round(data.get('intensity_score', 0), 6))
                row += 1
            
            # 7. 요약 시트
            summary_ws = self.output_workbook.create_sheet("요약")
            summary_ws.cell(row=1, column=1, value="구분")
            summary_ws.cell(row=1, column=2, value="개수")
            summary_ws.cell(row=2, column=1, value="전체 종목 수")
            summary_ws.cell(row=2, column=2, value=getattr(self, 'total_stock_count', len(self.eps_scores)))
            summary_ws.cell(row=3, column=1, value="EPS 필터 통과 종목 수")
            summary_ws.cell(row=3, column=2, value=len(self.eps_top_100))
            summary_ws.cell(row=4, column=1, value="최종 선정 종목 수")
            summary_ws.cell(row=4, column=2, value=len(self.final_top_50))
            summary_ws.cell(row=5, column=1, value="1개월 외국인 수급 상위 종목 수")
            summary_ws.cell(row=5, column=2, value=len(self.one_month_top_10))
            summary_ws.cell(row=6, column=1, value="2개월 외국인 수급 상위 종목 수")
            summary_ws.cell(row=6, column=2, value=len(self.two_month_top_10))
            summary_ws.cell(row=7, column=1, value="최종 비중 계산 종목 수")
            summary_ws.cell(row=7, column=2, value=len(self.final_weights))
            summary_ws.cell(row=8, column=1, value="총 선정 종목 수 (중복 포함)")
            summary_ws.cell(row=8, column=2, value=self.total_selection_count)
            
            # 파일 저장 (UTF-8 인코딩)
            self.output_workbook.save(self.output_excel_path)
            print(f"전체 종목 결과 Excel 파일 저장 완료: {self.output_excel_path}")
            
            return True
            
        except Exception as e:
            print(f"결과 Excel 파일 생성 실패: {e}")
            return False
    
    def run_full_stock_system(self, use_market_cap=True):
        """전체 종목 지수 리밸런싱 시스템 실행"""
        start_time = time.time()
        
        if not self.load_source_excel_file():
            return False
        
        print("=" * 80)
        print("DeepSearch 외인수급Top20 지수 (PR) 구성종목 선정 시스템 시작")
        print("영문명: DeepSearch Net Foreign BuyingTop20 Index PR")
        cap_type = "시가총액" if use_market_cap else "유동시가총액"
        print(f"사용 데이터: {cap_type}")
        print("=" * 80)
        
        # 1. 데이터 시트 찾기
        sheets = self.find_data_sheets(use_market_cap)
        if not sheets:
            return False
        
        # 2. 전체 종목 데이터 파싱
        eps_data, total_stock_count = self.parse_data(sheets.get('eps_sheet', ''), "eps")
        foreign_data, _ = self.parse_data(sheets.get('foreign_sheet', ''), "foreign")
        market_cap_data, _ = self.parse_data(sheets.get('market_cap_sheet', ''), "market_cap")
        
        # 전체 종목 수 저장 (원본 엑셀에서 추출한 종목코드 수)
        self.total_stock_count = total_stock_count
        
        if not eps_data or not foreign_data or not market_cap_data:
            print("필요한 데이터가 부족합니다.")
            return False
        
        # 3. EPS 필터 전체 종목 적용
        eps_filtered_stocks = self.apply_eps_filter(eps_data)
        if not eps_filtered_stocks:
            return False
        
        # 4. 외국인 수급강도 지표 전체 종목 계산
        final_stocks = self.calculate_foreign_intensity(eps_filtered_stocks, foreign_data, market_cap_data)
        if not final_stocks:
            return False
        
        # 5. 1개월과 2개월 외국인 수급 상위 10종목 계산
        one_month_top_10, two_month_top_10 = self.calculate_monthly_foreign_intensity(final_stocks, foreign_data, market_cap_data)
        if not one_month_top_10 or not two_month_top_10:
            return False
        
        # 6. 최종 비중 계산
        final_weights = self.calculate_final_weights()
        if not final_weights:
            return False
        
        # 7. 결과 Excel 파일 생성
        if not self.create_result_excel_full_stocks(self.final_top_50):
            print("결과 Excel 파일 생성 실패")
            return False
        
        end_time = time.time()
        execution_time = end_time - start_time
        
        print("=" * 80)
        print("DeepSearch 외인수급Top20 지수 (PR) 구성종목 선정 시스템 완료!")
        print("영문명: DeepSearch Net Foreign BuyingTop20 Index PR")
        print(f"- 전체 종목 수: {getattr(self, 'total_stock_count', len(self.eps_scores))}")
        print(f"- EPS 필터 통과 종목 수: {len(self.eps_top_100)}")
        print(f"- 최종 선정 종목 수: {len(self.final_top_50)}")
        print(f"- 1개월 외국인 수급 상위 종목 수: {len(self.one_month_top_10)}")
        print(f"- 2개월 외국인 수급 상위 종목 수: {len(self.two_month_top_10)}")
        print(f"- 최종 비중 계산 종목 수: {len(self.final_weights)}")
        print(f"- 총 선정 종목 수 (중복 포함): {self.total_selection_count}")
        print(f"- 실행 시간: {execution_time:.2f}초")
        print(f"- 결과 파일: {self.output_excel_path}")
        print("=" * 80)
        
        return True

class MonthlyRebalancingScheduler:
    """매달 리밸런싱 자동화 시스템"""
    
    def __init__(self, base_directory="excel_data"):
        self.base_directory = base_directory
        self.file_prefix = "deepsearch_net_foreign_buying_top20_index_raw_data_"
        self.result_prefix = "deepsearch_foreign_buying_top20_index_result_"
    
    def copy_file_with_custom_date(self, source_file, target_date):
        """사용자 지정 날짜로 파일 복사"""
        try:
            target_date_str = target_date.strftime('%Y%m%d')
            new_filename = f"{self.file_prefix}{target_date_str}.xlsx"
            source_path = os.path.join(self.base_directory, source_file)
            target_path = os.path.join(self.base_directory, new_filename)
            
            shutil.copy2(source_path, target_path)
            print(f"파일 복사 완료: {source_file} → {new_filename}")
            
            return new_filename, target_date
            
        except Exception as e:
            print(f"파일 복사 중 오류 발생: {e}")
            return None, None
    
    def update_dates_in_excel(self, filename, b5_value, b6_value):
        """Excel 파일 내의 날짜들을 사용자 입력값으로 업데이트 (전체 시트 순환)"""
        try:
            file_path = os.path.join(self.base_directory, filename)
            workbook = load_workbook(file_path, data_only=True)
            
            # 사용자 입력값 사용
            date_str = b5_value
            date_str_korean = b6_value
            
            # 전체 시트 개수 파악
            total_sheets = len(workbook.sheetnames)
            print(f"전체 시트 개수: {total_sheets}개")
            
            updated_sheets = 0
            
            # 모든 시트를 순환하면서 B5, B6 셀 업데이트 (첫 번째 시트부터 순서대로)
            for i, sheet_name in enumerate(workbook.sheetnames, 1):
                try:
                    sheet = workbook[sheet_name]
                    print(f"   - [{i}/{total_sheets}] {sheet_name} 시트 확인 중...")
                    
                    # B5 셀 확인 및 업데이트
                    current_b5_value = sheet['B5'].value
                    if current_b5_value:
                        sheet['B5'] = date_str
                        print(f"     B5 셀 업데이트: {current_b5_value} → {date_str}")
                    
                    # B6 셀 확인 및 업데이트
                    current_b6_value = sheet['B6'].value
                    if current_b6_value:
                        sheet['B6'] = date_str_korean
                        print(f"     B6 셀 업데이트: {current_b6_value} → {date_str_korean}")
                    
                    if current_b5_value or current_b6_value:
                        updated_sheets += 1
                        print(f"   {sheet_name} 시트 날짜 업데이트 완료")
                    else:
                        print(f"   {sheet_name} 시트는 날짜 셀이 비어있음")
                        
                except Exception as e:
                    print(f"   {sheet_name} 시트 처리 중 오류: {e}")
            
            # 파일 저장 (UTF-8 인코딩)
            workbook.save(file_path)
            print(f"Excel 파일 날짜 업데이트 완료: {filename}")
            print(f"업데이트된 시트 수: {updated_sheets}개")
            
            return True
            
        except Exception as e:
            print(f"Excel 파일 날짜 업데이트 중 오류 발생: {e}")
            return False
    
    def open_excel_and_refresh_data(self, filename, automation_mode="macro"):
        """Excel 파일을 열어서 Quantiwise refresh 후 저장"""
        try:
            import os
            
            file_path = os.path.join(self.base_directory, filename)
            
            if automation_mode == "macro":
                print(f"Excel 매크로 자동화 모드: {filename}")
                
                try:
                    import win32com.client as win32
                    
                    # 절대 경로로 변환
                    absolute_file_path = os.path.abspath(file_path)
                    print(f"Excel 파일 열기: {absolute_file_path}")
                    
                    # 파일 존재 여부 확인
                    if not os.path.exists(absolute_file_path):
                        print(f"파일이 존재하지 않습니다: {absolute_file_path}")
                        return False
                    
                    # Excel 인스턴스 생성
                    excel = win32.Dispatch("Excel.Application")
                    excel.Visible = True
                    
                    # Excel 파일 열기
                    workbook = excel.Workbooks.Open(absolute_file_path)
                    
                    print("Quantiwise refresh 매크로 실행 중...")
                    
                    # 전체 시트 개수 파악 및 순환 처리
                    total_sheets = workbook.Worksheets.Count
                    print(f"전체 시트 개수: {total_sheets}개")
                    
                    # 첫 번째 시트로 먼저 이동
                    print("첫 번째 시트로 이동 중...")
                    first_sheet = workbook.Worksheets(1)
                    first_sheet.Activate()
                    print(f"첫 번째 시트 활성화: {first_sheet.Name}")
                    
                    refresh_success_count = 0
                    processed_sheets = 0
                    
                    # 모든 시트를 순환하면서 refresh 버튼이 있는 시트만 처리 (첫 번째 시트부터 순서대로)
                    for i in range(1, total_sheets + 1):
                        try:
                            worksheet = workbook.Worksheets(i)
                            sheet_name = worksheet.Name
                            
                            print(f"   - [{i}/{total_sheets}] {sheet_name} 시트 확인 중...")
                            
                            # 각 시트를 활성화
                            worksheet.Activate()
                            print(f"   {sheet_name} 시트 활성화 완료")
                            
                            # A1 셀에 refresh 버튼이 있는지 확인 (퀀티와이즈 가이드 기반)
                            try:
                                # A1 셀 직접 접근
                                a1_cell = worksheet.Range("A1")
                                
                                # A1 셀에 hyperlink가 있는지 확인
                                hyperlink_count = a1_cell.Hyperlinks.Count
                                
                                if hyperlink_count > 0:
                                    # A1 셀 값이 "Refresh"인지 확인
                                    cell_value = str(a1_cell.Value).strip()
                                    if "Refresh" in cell_value:
                                        print(f"   {sheet_name} 시트 refresh 실행 중...")
                                        
                                        # 퀀티와이즈 가이드에 따른 Refresh 버튼 실행
                                        # Range("A1").Select
                                        worksheet.Range("A1").Select()
                                        
                                        # Selection.Hyperlinks(1).Follow NewWindow:=False, AddHistory:=True
                                        excel.Selection.Hyperlinks(1).Follow(NewWindow=False, AddHistory=True)
                                        
                                        # 데이터 로딩 대기
                                        time.sleep(5)
                                        
                                        print(f"   {sheet_name} 시트 refresh 완료")
                                        refresh_success_count += 1
                                        processed_sheets += 1
                                    else:
                                        print(f"   {sheet_name} 시트는 refresh 대상이 아닙니다 (A1 값: {cell_value})")
                                else:
                                    print(f"   {sheet_name} 시트는 refresh 버튼이 없습니다")
                                    
                            except Exception as hyperlink_error:
                                print(f"   {sheet_name} 시트 refresh 버튼 확인 실패: {hyperlink_error}")
                                
                        except Exception as e:
                            print(f"   {sheet_name} 시트 처리 중 오류: {e}")
                    
                    print(f"처리 결과: {processed_sheets}개 시트 중 {refresh_success_count}개 성공")
                    
                    # 매크로 실행 결과 확인
                    if processed_sheets == 0:
                        print("refresh 대상 시트를 찾을 수 없습니다.")
                        workbook.Close()
                        excel.Quit()
                        return False
                    elif refresh_success_count == 0:
                        print("모든 refresh 대상 시트에서 실패했습니다.")
                        workbook.Close()
                        excel.Quit()
                        return False
                    elif refresh_success_count < processed_sheets:
                        print(f"{refresh_success_count}/{processed_sheets} 시트에서만 refresh 성공했습니다.")
                        workbook.Close()
                        excel.Quit()
                        return False
                    
                    # 파일 저장
                    print("파일 저장 중...")
                    workbook.Save()
                    
                    # Excel 닫기
                    workbook.Close()
                    excel.Quit()
                    
                    print("Excel 매크로 자동화 완료")
                    
                except ImportError:
                    print("pywin32 라이브러리가 설치되지 않았습니다.")
                    print("pip install pywin32 명령으로 설치 후 다시 시도하세요.")
                    return False
                except Exception as e:
                    print(f"Excel 매크로 자동화 실패: {e}")
                    return False
            
            print(f"Excel 파일 처리 완료: {filename}")
            return True
            
        except Exception as e:
            print(f"Excel 파일 열기 중 오류 발생: {e}")
            return False
    
    def run_analysis(self, filename, use_market_cap=True):
        """업데이트된 파일로 분석 실행"""
        try:
            input_file = os.path.join(self.base_directory, filename)
            
            # 결과 파일명 생성
            date_str = filename.replace(self.file_prefix, '').replace('.xlsx', '')
            if use_market_cap:
                result_filename = f"{self.result_prefix}{date_str}.xlsx"
            else:
                result_filename = f"{self.result_prefix}ff_{date_str}.xlsx"
            output_file = os.path.join(self.base_directory, result_filename)
            
            print(f"분석 시작: {filename}")
            print(f"사용 데이터: {'시가총액' if use_market_cap else '유동시가총액'}")
            print(f"결과 파일: {result_filename}")
            
            # DeepSearch 시스템 실행
            system = DeepSearchForeignBuyingTop20IndexSystem(input_file, output_file)
            success = system.run_full_stock_system(use_market_cap)
            
            if success:
                print(f"분석 완료: {result_filename}")
                return True
            else:
                print(f"분석 실패: {filename}")
                return False
                
        except Exception as e:
            print(f"분석 실행 중 오류 발생: {e}")
            return False

def main():
    """메인 실행 함수 - 매달 리밸런싱 자동화"""
    print("=" * 80)
    print("DeepSearch 외인수급Top20 지수 매달 리밸런싱 시작")
    print("영문명: DeepSearch Net Foreign BuyingTop20 Index PR")
    print("=" * 80)
    
    # 매달 리밸런싱 스케줄러 초기화
    scheduler = MonthlyRebalancingScheduler()
    
    # 사용자 입력 받기
    try:
        # datetime 모듈 import
        from datetime import datetime, timedelta
        
        print("\n사용자 입력:")
        
        # 1. 기존 파일 날짜 입력
        print("1. 기존 raw_data 파일의 날짜를 입력하세요 (YYYY-MM-DD 형식)")
        print("예: 2025-08-31")
        existing_date_input = input("기존 파일 날짜: ").strip()
        
        # 2. 시가총액 타입 선택
        print("\n2. 시가총액 타입을 선택하세요:")
        print("   1) 시가총액 사용")
        print("   2) 유동시가총액 사용")
        cap_choice = input("선택 (1 또는 2): ").strip()
        
        if cap_choice == "1":
            use_market_cap = True
            cap_type_name = "시가총액"
        elif cap_choice == "2":
            use_market_cap = False
            cap_type_name = "유동시가총액"
        else:
            print("잘못된 선택입니다. 1 또는 2를 입력해주세요.")
            return
        
        # 3. 작업 방식 선택
        print(f"\n3. 작업 방식을 선택하세요 ({cap_type_name} 사용):")
        print("   1) 새 엑셀 파일 생성 (기존 파일 복사 후 날짜 업데이트)")
        print("   2) 기존 엑셀 파일 사용 (이미 생성된 파일 활용)")
        choice = input("선택 (1 또는 2): ").strip()
        
        if choice == "1":
            # 새 파일 생성 모드
            print("\n새 파일 생성 모드:")
            print("새로 생성할 raw_data 파일의 날짜를 입력하세요 (YYYY-MM-DD 형식)")
            print("예: 2025-09-30")
            new_date_input = input("새 파일 날짜: ").strip()
            new_date = datetime.strptime(new_date_input, '%Y-%m-%d')
            create_new_file = True
        elif choice == "2":
            # 기존 파일 사용 모드
            print("\n기존 파일 사용 모드:")
            print("사용할 raw_data 파일의 날짜를 입력하세요 (YYYY-MM-DD 형식)")
            print("예: 2025-09-30")
            new_date_input = input("기존 파일 날짜: ").strip()
            new_date = datetime.strptime(new_date_input, '%Y-%m-%d')
            create_new_file = False
        else:
            print("잘못된 선택입니다. 1 또는 2를 입력해주세요.")
            return
        
        # 새 파일 날짜를 YYYYMMDD 형식으로 변환하여 B6 셀 값으로 사용
        b6_value_input = new_date.strftime('%Y%m%d')
        
        print(f"\n입력된 정보:")
        print(f"   기존 파일 날짜: {existing_date_input}")
        print(f"   시가총액 타입: {cap_type_name}")
        print(f"   작업 방식: {'새 파일 생성' if create_new_file else '기존 파일 사용'}")
        print(f"   대상 파일 날짜: {new_date_input}")
        print(f"   B6 셀 값 (자동 변환): {b6_value_input}")
        
        # B5 셀 값은 B6 기준으로 1년 전 날짜 자동 계산
        b6_date = datetime.strptime(b6_value_input, '%Y%m%d')
        b5_date = b6_date - timedelta(days=365)  # 1년 전
        b5_value_input = b5_date.strftime('%Y%m%d')
        print(f"   B5 셀 값 (자동 계산): {b5_value_input} (B6 기준 1년 전)")
        
        # 날짜 파싱
        existing_date = datetime.strptime(existing_date_input, '%Y-%m-%d')
        
        print(f"\n최종 설정:")
        print(f"   기존 파일 날짜: {existing_date.strftime('%Y년 %m월 %d일')}")
        print(f"   시가총액 타입: {cap_type_name}")
        print(f"   작업 방식: {'새 파일 생성' if create_new_file else '기존 파일 사용'}")
        print(f"   대상 파일 날짜: {new_date.strftime('%Y년 %m월 %d일')}")
        print(f"   B5 셀 값: {b5_value_input}")
        print(f"   B6 셀 값: {b6_value_input}")
        print("리밸런싱을 시작합니다...")
        
    except ValueError as e:
        print(f"입력 형식 오류가 발생했습니다: {e}")
        print("올바른 날짜 형식으로 다시 입력해주세요.")
        return
    except KeyboardInterrupt:
        print("\n사용자가 취소했습니다.")
        return
    
    # 리밸런싱 프로세스 시작
    start_time = time.time()
    new_filename = None  # 새로 생성된 파일명 추적
    
    try:
        if create_new_file:
            # 새 파일 생성 모드
            print(f"\n새 파일 생성 모드:")
            
            # 1. 기존 파일 찾기
            print(f"기존 파일({existing_date.strftime('%Y%m%d')}) 검색 중...")
            existing_filename = f"{scheduler.file_prefix}{existing_date.strftime('%Y%m%d')}.xlsx"
            existing_path = os.path.join(scheduler.base_directory, existing_filename)
            
            if not os.path.exists(existing_path):
                print(f"해당 날짜의 파일을 찾을 수 없습니다: {existing_filename}")
                print("사용 가능한 파일 목록:")
                files = [f for f in os.listdir(scheduler.base_directory) 
                        if f.startswith(scheduler.file_prefix) and f.endswith('.xlsx')]
                for file in sorted(files):
                    print(f"  - {file}")
                return
            
            print(f"기존 파일 발견: {existing_filename}")
            
            # 2. 새 파일로 복사
            print("새 파일로 복사 중...")
            new_filename, new_date = scheduler.copy_file_with_custom_date(existing_filename, new_date)
            if not new_filename:
                return
            
            # 3. Excel 파일 내 날짜 업데이트
        else:
            # 기존 파일 사용 모드
            print(f"\n기존 파일 사용 모드:")
            
            # 대상 파일 찾기
            target_filename = f"{scheduler.file_prefix}{new_date.strftime('%Y%m%d')}.xlsx"
            target_path = os.path.join(scheduler.base_directory, target_filename)
            
            if not os.path.exists(target_path):
                print(f"해당 날짜의 파일을 찾을 수 없습니다: {target_filename}")
                print("사용 가능한 파일 목록:")
                files = [f for f in os.listdir(scheduler.base_directory) 
                        if f.startswith(scheduler.file_prefix) and f.endswith('.xlsx')]
                for file in sorted(files):
                    print(f"  - {file}")
                return
            
            print(f"기존 파일 발견: {target_filename}")
            new_filename = target_filename
            existing_filename = target_filename  # 기존 파일 사용 모드에서는 같은 파일
        
        # 4. Excel 파일 내 날짜 업데이트 (새 파일 생성 모드에서만)
        if create_new_file:
            print("Excel 파일 내 날짜 업데이트 중...")
            if not scheduler.update_dates_in_excel(new_filename, b5_value_input, b6_value_input):
                raise Exception("Excel 파일 날짜 업데이트 실패")
        else:
            print("기존 파일 사용 모드: 날짜 업데이트 건너뜀")
        
        # 4. Excel 파일 열기 및 Quantiwise refresh (새 파일 생성 모드에서만)
        if create_new_file:
            print("Excel 파일 열기 및 데이터 새로고침 중...")
            if not scheduler.open_excel_and_refresh_data(new_filename, "macro"):
                raise Exception("Excel 파일 refresh 실패")
        else:
            print("기존 파일 사용 모드: Excel refresh 건너뜀")
        
        # 5. 분석 실행
        print("데이터 분석 실행 중...")
        if not scheduler.run_analysis(new_filename, use_market_cap):
            raise Exception("데이터 분석 실패")
        
        # 6. 완료 메시지
        end_time = time.time()
        execution_time = end_time - start_time
        
        print("=" * 80)
        print("DeepSearch 외인수급Top20 지수 매달 리밸런싱 완료!")
        print(f"- 기존 파일: {existing_filename}")
        print(f"- 새 파일: {new_filename}")
        print(f"- 결과 파일: {scheduler.result_prefix}{new_date.strftime('%Y%m%d')}.xlsx")
        print(f"- 실행 시간: {execution_time:.2f}초")
        print("=" * 80)
        
    except Exception as e:
        # 에러 발생 시 새로 생성된 파일만 삭제 (기존 파일 사용 모드에서는 삭제하지 않음)
        if new_filename and create_new_file:  # 새 파일 생성 모드에서만 삭제
            try:
                new_file_path = os.path.join(scheduler.base_directory, new_filename)
                if os.path.exists(new_file_path):
                    os.remove(new_file_path)
                    print(f"에러 발생으로 인한 파일 정리: {new_filename} 삭제 완료")
            except Exception as cleanup_error:
                print(f"파일 삭제 중 오류 발생: {cleanup_error}")
        
        print(f"리밸런싱 프로세스 실패: {e}")
        print("=" * 80)

if __name__ == "__main__":
    main()
